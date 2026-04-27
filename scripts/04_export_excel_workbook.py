from __future__ import annotations

import csv
import datetime as dt
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
META_DIR = ROOT / "02_元数据"
OUTPUT_DIR = ROOT / "05_输出成果"
WORKBOOK_PATH = META_DIR / "知识库管理工作簿.xlsx"

SOURCES = [
    ("问题评测表", META_DIR / "问题评测表.csv"),
    ("人工核验表", META_DIR / "人工核验表.csv"),
    ("政策资料台账", META_DIR / "政策资料台账.csv"),
    ("规则清单", META_DIR / "规则清单.csv"),
    ("政策关联关系", META_DIR / "政策关联关系表.csv"),
    ("来源清单", META_DIR / "来源清单.csv"),
    ("地区覆盖清单", META_DIR / "地区覆盖清单.csv"),
    ("待采集链接", META_DIR / "待采集链接.csv"),
]


def latest_retrieval_result() -> Path | None:
    files = sorted(OUTPUT_DIR.glob("检索评测结果_*.csv"), key=lambda p: p.name, reverse=True)
    return files[0] if files else None


def latest_search_result() -> Path | None:
    files = sorted(OUTPUT_DIR.glob("全文检索结果_*.csv"), key=lambda p: p.name, reverse=True)
    return files[0] if files else None


def latest_title_audit() -> Path | None:
    files = sorted(OUTPUT_DIR.glob("标题清洗建议_*.csv"), key=lambda p: p.name, reverse=True)
    return files[0] if files else None


def read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def fit_width(value: str) -> int:
    # Chinese characters are wider in Excel, so count them a little heavier.
    width = 0
    for char in str(value):
        width += 2 if ord(char) > 127 else 1
    return min(max(width + 2, 10), 42)


def write_table_sheet(wb: Workbook, sheet_name: str, path: Path) -> None:
    ws = wb.create_sheet(title=sheet_name)
    headers, rows = read_csv(path)
    if not headers:
        ws.append(["提示"])
        ws.append([f"{path.name} 为空"])
        return

    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, header in enumerate(headers, start=1):
        values = [header]
        values.extend(row[col_idx - 1] for row in rows[:80] if col_idx - 1 < len(row))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(fit_width(v) for v in values)


def write_readme_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(title="使用说明", index=0)
    rows = [
        ["项目", "说明"],
        ["生成时间", dt.datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["文件用途", "把知识库关键CSV汇总为一个本地Excel工作簿，方便人工查看、筛选和核验。"],
        ["重要提示", "当前脚本仍以CSV为主数据源。若直接修改Excel，后续需要再同步回CSV。"],
        ["建议流程", "先在Excel中筛选和核验，再把确认后的修正同步回对应CSV。"],
        ["问题评测表", "用于维护标准问题、标准依据文件、标准答案要点和检索评测结果。"],
        ["人工核验表", "用于逐份核验标题、日期、部门、地区、标签、有效状态和正文完整性。"],
        ["政策资料台账", "用于查看已采集入库资料。"],
        ["规则清单", "从台账抽取规则名称、发布机构、文号、省份和原文链接，方便对照规则库目录。"],
        ["政策关联关系", "仅记录新政策明文引用旧政策的关系，用于搜索页库内跳转和版本脉络核验。"],
        ["来源清单", "用于控制自动采集的白名单来源。"],
        ["地区覆盖清单", "用于查看省级、直辖市、副省级城市和电网经营区覆盖情况。"],
        ["待采集链接", "用于人工确认后交给采集脚本下载入库。"],
    ]
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2F0D9")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_workbook() -> Path:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    write_readme_sheet(wb)
    for sheet_name, path in SOURCES:
        if path.exists():
            write_table_sheet(wb, sheet_name, path)

    result_path = latest_retrieval_result()
    if result_path:
        write_table_sheet(wb, "检索评测结果", result_path)

    search_path = latest_search_result()
    if search_path:
        write_table_sheet(wb, "全文检索结果", search_path)

    title_audit_path = latest_title_audit()
    if title_audit_path:
        write_table_sheet(wb, "标题清洗建议", title_audit_path)

    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(WORKBOOK_PATH)
        return WORKBOOK_PATH
    except PermissionError:
        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback_path = META_DIR / f"知识库管理工作簿_更新版_{timestamp}.xlsx"
        wb.save(fallback_path)
        return fallback_path


def self_check(path: Path) -> list[str]:
    errors: list[str] = []
    if not path.exists():
        return [f"未生成文件：{path}"]

    wb = load_workbook(path, read_only=True, data_only=True)
    expected_sheets = ["使用说明"] + [name for name, source in SOURCES if source.exists()]
    if latest_retrieval_result():
        expected_sheets.append("检索评测结果")
    if latest_search_result():
        expected_sheets.append("全文检索结果")
    if latest_title_audit():
        expected_sheets.append("标题清洗建议")

    for sheet_name in expected_sheets:
        if sheet_name not in wb.sheetnames:
            errors.append(f"缺少工作表：{sheet_name}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 1 or ws.max_column < 1:
            errors.append(f"工作表为空：{sheet_name}")
    wb.close()
    return errors


def main() -> int:
    path = build_workbook()
    errors = self_check(path)
    if errors:
        print("Excel工作簿自检失败：")
        for error in errors:
            print(f"- {error}")
        return 1

    wb = load_workbook(path, read_only=True)
    sheet_names = ", ".join(wb.sheetnames)
    wb.close()
    print(f"Excel工作簿已生成：{path}")
    print(f"工作表：{sheet_names}")
    print("自检通过")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
