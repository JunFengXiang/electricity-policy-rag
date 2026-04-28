"""将人工编辑后的 Excel 工作表同步回 CSV 元数据表。

同步前会做表头校验并备份原 CSV，避免 Excel 手工维护时把流水线依赖的字段
误删或改名。
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import shutil
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
META_DIR = ROOT / "02_元数据"
WORKBOOK_PATH = META_DIR / "知识库管理工作簿.xlsx"
BACKUP_DIR = META_DIR / "备份"

SYNC_TARGETS = {
    "问题评测表": META_DIR / "问题评测表.csv",
    "人工核验表": META_DIR / "人工核验表.csv",
    "政策资料台账": META_DIR / "政策资料台账.csv",
    "来源清单": META_DIR / "来源清单.csv",
    "地区覆盖清单": META_DIR / "地区覆盖清单.csv",
    "待采集链接": META_DIR / "待采集链接.csv",
}

KEY_FIELDS = {
    "问题评测表": "评测编号",
    "人工核验表": "核验编号",
    "政策资料台账": "资料编号",
    "来源清单": "来源编号",
    "地区覆盖清单": "地区编号",
    "待采集链接": "url",
}


def read_csv_header(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        return next(reader, [])


def sheet_to_rows(workbook_path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, str]]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作簿缺少工作表：{sheet_name}")

        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not raw_rows:
        return [], []

    headers = [str(value).strip() if value is not None else "" for value in raw_rows[0]]
    headers = [header for header in headers if header]
    rows: list[dict[str, str]] = []

    for raw_row in raw_rows[1:]:
        values = []
        for value in raw_row[: len(headers)]:
            if value is None:
                values.append("")
            else:
                values.append(str(value).strip())

        if not any(values):
            continue
        rows.append(dict(zip(headers, values, strict=False)))

    return headers, rows


def validate_sheet(sheet_name: str, target_path: Path, headers: list[str], rows: list[dict[str, str]]) -> list[str]:
    """同步前校验表头，防止 Excel 中的列被误删或改名后覆盖 CSV。"""
    errors: list[str] = []
    if not target_path.exists():
        errors.append(f"{sheet_name}: 目标CSV不存在：{target_path}")
        return errors

    csv_headers = read_csv_header(target_path)
    if headers != csv_headers:
        missing = [header for header in csv_headers if header not in headers]
        extra = [header for header in headers if header not in csv_headers]
        errors.append(
            f"{sheet_name}: 表头与CSV不一致；缺少字段={';'.join(missing) or '无'}；额外字段={';'.join(extra) or '无'}"
        )

    key_field = KEY_FIELDS.get(sheet_name)
    if key_field and key_field in headers:
        seen: set[str] = set()
        for index, row in enumerate(rows, start=2):
            key = row.get(key_field, "")
            if not key:
                errors.append(f"{sheet_name}: 第{index}行缺少主键字段 {key_field}")
                continue
            if key in seen:
                errors.append(f"{sheet_name}: 主键字段 {key_field} 重复：{key}")
            seen.add(key)

    return errors


def write_csv(path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def backup_csv(path: Path, timestamp: str) -> Path:
    """每次写回 CSV 前都留一份备份，方便人工误操作后回滚。"""
    target_dir = BACKUP_DIR / timestamp
    target_dir.mkdir(parents=True, exist_ok=True)
    backup_path = target_dir / path.name
    shutil.copy2(path, backup_path)
    return backup_path


def selected_targets(sheet_names: list[str] | None) -> dict[str, Path]:
    if not sheet_names:
        return SYNC_TARGETS
    unknown = [name for name in sheet_names if name not in SYNC_TARGETS]
    if unknown:
        raise ValueError(f"未知工作表：{';'.join(unknown)}")
    return {name: SYNC_TARGETS[name] for name in sheet_names}


def run_sync(workbook_path: Path, sheet_names: list[str] | None, write: bool) -> int:
    """执行 Excel 到 CSV 的同步；未加 --write 时只做预检。"""
    if not workbook_path.exists():
        print(f"未找到Excel工作簿：{workbook_path}")
        return 1

    targets = selected_targets(sheet_names)
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    all_errors: list[str] = []
    loaded: list[tuple[str, Path, list[str], list[dict[str, str]]]] = []

    for sheet_name, target_path in targets.items():
        headers, rows = sheet_to_rows(workbook_path, sheet_name)
        errors = validate_sheet(sheet_name, target_path, headers, rows)
        if errors:
            all_errors.extend(errors)
        loaded.append((sheet_name, target_path, headers, rows))

    if all_errors:
        print("同步前校验失败：")
        for error in all_errors:
            print(f"- {error}")
        return 1

    mode = "写入" if write else "预检查"
    print(f"Excel同步{mode}通过：{len(loaded)} 个工作表")
    for sheet_name, target_path, _, rows in loaded:
        print(f"- {sheet_name}: {len(rows)} 行 -> {target_path}")

    if not write:
        print("当前为dry-run，未写入CSV。需要写回时追加 --write")
        return 0

    for sheet_name, target_path, headers, rows in loaded:
        backup_path = backup_csv(target_path, timestamp)
        write_csv(target_path, headers, rows)
        print(f"- 已写回 {sheet_name}，备份：{backup_path}")

    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync selected sheets from the Excel workbook back to CSV files.")
    parser.add_argument("--workbook", default=str(WORKBOOK_PATH), help="Excel workbook path.")
    parser.add_argument("--sheet", action="append", help="Sheet name to sync. Can be used multiple times.")
    parser.add_argument("--write", action="store_true", help="Write CSV files. Without this flag the script only validates.")
    args = parser.parse_args()

    try:
        return run_sync(Path(args.workbook), args.sheet, args.write)
    except Exception as exc:
        print(f"同步失败：{exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
